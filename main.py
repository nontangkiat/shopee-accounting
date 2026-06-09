from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
import anthropic, base64, json, os, uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

PROMPT = """Extract Shopee purchase data from this screenshot. Return ONLY raw JSON, no markdown.

{"orders":[{"order_id":"","date":"","shop_name":"","item_name":"","quantity":1,"unit_price":0,"total_price":0,"net_total":0,"shipping_fee":0,"discount":0,"payment_method":"","status":""}]}

- Extract every item visible
- Prices as numbers in THB (no symbols)
- Missing fields: "" or 0
- net_total = actual amount paid"""

HTML = """<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Shopee → บัญชี</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Sarabun', sans-serif; background: #f5f5f5; color: #333; min-height: 100vh; }
  .header { background: #ee4d2d; padding: 16px 24px; display: flex; align-items: center; gap: 12px; }
  .header h1 { color: white; font-size: 18px; font-weight: 700; }
  .header span { color: rgba(255,255,255,0.8); font-size: 13px; }
  .container { max-width: 640px; margin: 40px auto; padding: 0 20px; }
  .card { background: white; border-radius: 12px; padding: 28px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }
  .dropzone { border: 2px dashed #ddd; border-radius: 10px; padding: 44px 20px; text-align: center; cursor: pointer; transition: all 0.15s; }
  .dropzone:hover, .dropzone.over { border-color: #ee4d2d; background: #fff5f3; }
  .dropzone .icon { font-size: 40px; margin-bottom: 10px; }
  .dropzone p { font-size: 15px; font-weight: 600; margin-bottom: 4px; }
  .dropzone small { color: #999; font-size: 12px; }
  #fileInput { display: none; }
  .file-list { margin-top: 16px; display: flex; flex-direction: column; gap: 8px; }
  .file-item { display: flex; align-items: center; gap: 10px; padding: 10px 14px; background: #f9f9f9; border-radius: 8px; border: 1px solid #eee; }
  .file-item img { width: 40px; height: 50px; object-fit: cover; border-radius: 4px; border: 1px solid #ddd; }
  .file-item .name { flex: 1; font-size: 13px; color: #555; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .file-item button { background: none; border: none; color: #bbb; cursor: pointer; font-size: 16px; }
  .file-item button:hover { color: #ee4d2d; }
  .btn { width: 100%; padding: 14px; border-radius: 10px; border: none; font-size: 15px; font-weight: 700; cursor: pointer; margin-top: 18px; }
  .btn-primary { background: linear-gradient(135deg, #ee4d2d, #f89c1c); color: white; }
  .btn-primary:disabled { background: #ddd; color: #aaa; cursor: not-allowed; }
  .status { margin-top: 14px; padding: 12px 16px; border-radius: 8px; font-size: 13px; display: none; }
  .status.loading { background: #e8f4fd; color: #1a73e8; border: 1px solid #c0d9f5; display: block; }
  .status.error { background: #fde8e8; color: #c0392b; border: 1px solid #f5c0c0; display: block; word-break: break-all; }
  .status.success { background: #e8fdf0; color: #1a7340; border: 1px solid #b0e8c8; display: block; }
  .tip { font-size: 12px; color: #999; text-align: center; margin-top: 10px; }
</style>
</head>
<body>
<div class="header">
  <div style="font-size:24px">🛍</div>
  <div><h1>Shopee → บัญชี</h1><span>อัปโหลด screenshot → ได้ Excel พร้อมรูปหลักฐาน</span></div>
</div>
<div class="container">
  <div class="card">
    <div class="dropzone" id="dropzone" onclick="document.getElementById('fileInput').click()">
      <div class="icon">📂</div>
      <p>วาง หรือ คลิกเพื่อเลือกรูป Shopee</p>
      <small>รองรับ .jpg .png — หลายไฟล์ได้</small>
    </div>
    <input type="file" id="fileInput" accept="image/*" multiple>
    <div class="file-list" id="fileList"></div>
    <div class="status" id="status"></div>
    <button class="btn btn-primary" id="processBtn" disabled onclick="process()">🔍 วิเคราะห์และสร้าง Excel</button>
    <p class="tip">รูปถูกฝังใน Excel เป็นหลักฐาน · ข้อมูลไม่ถูกบันทึกบน server</p>
  </div>
</div>
<script>
let files = [];
const dropzone = document.getElementById('dropzone');
const fileInput = document.getElementById('fileInput');
const fileList = document.getElementById('fileList');
const btn = document.getElementById('processBtn');
const statusEl = document.getElementById('status');

dropzone.addEventListener('dragover', e => { e.preventDefault(); dropzone.classList.add('over'); });
dropzone.addEventListener('dragleave', () => dropzone.classList.remove('over'));
dropzone.addEventListener('drop', e => { e.preventDefault(); dropzone.classList.remove('over'); addFiles(e.dataTransfer.files); });
fileInput.addEventListener('change', e => addFiles(e.target.files));

function addFiles(newFiles) {
  Array.from(newFiles).filter(f => f.type.startsWith('image/')).forEach(f => {
    files.push(f);
    const url = URL.createObjectURL(f);
    const div = document.createElement('div');
    div.className = 'file-item';
    div.dataset.name = f.name;
    div.innerHTML = `<img src="${url}"><span class="name">${f.name}</span><button onclick="removeFile('${f.name}',this)">✕</button>`;
    fileList.appendChild(div);
  });
  btn.disabled = files.length === 0;
}
function removeFile(name, el) {
  files = files.filter(f => f.name !== name);
  el.closest('.file-item').remove();
  btn.disabled = files.length === 0;
}
async function process() {
  btn.disabled = true;
  setStatus('loading', `⏳ กำลังวิเคราะห์ ${files.length} รูป...`);
  try {
    const form = new FormData();
    files.forEach(f => form.append('files', f));
    const res = await fetch('/process', { method: 'POST', body: form });
    if (!res.ok) { const t = await res.text(); throw new Error(`HTTP ${res.status}: ${t.slice(0,200)}`); }
    const blob = await res.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = `shopee_orders_${new Date().toISOString().slice(0,10)}.xlsx`;
    a.click();
    setStatus('success', `✅ สร้าง Excel สำเร็จ ${files.length} รูป — กำลัง download...`);
  } catch(err) { setStatus('error', 'Error: ' + err.message); }
  btn.disabled = false;
}
function setStatus(type, msg) {
  statusEl.className = 'status ' + type;
  statusEl.textContent = msg;
}
</script>
</body>
</html>"""

def extract_data(image_bytes: bytes, media_type: str) -> list:
    b64 = base64.standard_b64encode(image_bytes).decode()
    msg = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=1000,
        messages=[{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
            {"type": "text", "text": PROMPT}
        ]}]
    )
    text = msg.content[0].text
    s, e = text.index("{"), text.rindex("}")
    return json.loads(text[s:e+1]).get("orders", [])

def make_excel(all_orders: list, image_map: dict) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Shopee Orders"
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", start_color="EE4D2D")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", start_color="FFF8F7")
    headers = [("หลักฐาน",18),("Order ID",22),("วันที่",13),("ร้านค้า",26),
               ("ชื่อสินค้า",34),("จำนวน",8),("ราคา/หน่วย",12),("รวม",11),
               ("ยอดสุทธิ",11),("ค่าส่ง",9),("สถานะ",20)]
    for c, (h, w) in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = bdr
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28
    ROW_H = 300
    for ri, order in enumerate(all_orders, 2):
        ws.row_dimensions[ri].height = ROW_H
        row_fill = alt_fill if ri % 2 == 0 else None
        vals = ["", order.get("order_id",""), order.get("date",""),
                order.get("shop_name",""), order.get("item_name",""),
                order.get("quantity",0), order.get("unit_price",0),
                order.get("total_price",0), order.get("net_total",0),
                order.get("shipping_fee",0), order.get("status","")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(ri, c, v)
            cell.font = Font(name="Arial", size=10)
            cell.border = bdr
            if row_fill: cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True,
                horizontal="center" if c in (1,6) else "right" if c in (7,8,9,10) else "left")
        img_bytes = image_map.get(order.get("_file",""))
        if img_bytes:
            DISPLAY_H = 400
            with PILImage.open(io.BytesIO(img_bytes)) as im:
                im = im.convert("RGB")
                orig_w, orig_h = im.size
                ratio = DISPLAY_H / orig_h
                display_w = int(orig_w * ratio)
                buf = io.BytesIO()
                im.save(buf, "JPEG", quality=95)
                buf.seek(0)
            tmp = f"/tmp/img_{ri}.jpg"
            with open(tmp, "wb") as f: f.write(buf.read())
            xl_img = XLImage(tmp)
            xl_img.width = display_w
            xl_img.height = DISPLAY_H
            ws.add_image(xl_img, f"A{ri}")
    sr = len(all_orders) + 2
    ws.cell(sr, 4, "รวมทั้งหมด").font = Font(name="Arial", bold=True, size=11)
    ws.cell(sr, 8, f"=SUM(H2:H{sr-1})").font = Font(name="Arial", bold=True)
    ws.cell(sr, 9, f"=SUM(I2:I{sr-1})").font = Font(name="Arial", bold=True)
    sum_fill = PatternFill("solid", start_color="FFE5E0")
    for c in range(1, 12):
        ws.cell(sr, c).border = bdr
        ws.cell(sr, c).fill = sum_fill
    ws.freeze_panes = "B2"
    out = f"/tmp/shopee_{uuid.uuid4().hex[:8]}.xlsx"
    wb.save(out)
    return out

@app.get("/", response_class=HTMLResponse)
def index(): return HTML

@app.post("/process")
async def process(files: list[UploadFile] = File(...)):
    all_orders = []
    image_map = {}
    for f in files:
        data = await f.read()
        mt = f.content_type or "image/jpeg"
        if not mt.startswith("image/"): mt = "image/jpeg"
        orders = extract_data(data, mt)
        for o in orders: o["_file"] = f.filename
        image_map[f.filename] = data
        all_orders.extend(orders)
    path = make_excel(all_orders, image_map)
    return FileResponse(path, filename="shopee_orders.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/health")
def health(): return {"status": "ok"}
